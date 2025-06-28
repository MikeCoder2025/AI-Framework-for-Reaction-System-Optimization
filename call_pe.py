# -*- coding: utf-8 -*-
"""
Created on Tue Apr  7 11:53:08 2020

@author: p_aru
"""
import time

start_time = time.time()

import os
from matplotlib import pyplot as plt
import numpy as np
import pandas as pd
from scipy.optimize import minimize_scalar
from scipy.optimize import minimize
import scipy.stats as stats
from openpyxl import load_workbook
import random
import pyomo.environ as py
from closedloopmain import FIM_fun, MLE_pd, insilico_exp, obs_FIM_ed
from closedloopmain import km1, km2, km3
from closedloopmain import residual
from closedloopmain import obs_FIM, obs_COR, pred_error
from closedloopmain import tvalue_fun, mprob1, mprob2
from closedloopmain import optikm1, optikm2, optikm3
from closedloopmain import mbdoepp, initialisation0_mbdoepp, initialisation1_mbdoepp
from closedloopmain import initialisation0_mbdoemd_BF, initialisation1_mbdoemd_BF, mbdoemd_BF
from closedloopmain_modified import FIM_fun_ed
from closedloopmain import optimize_conversion
from closedloopmain import optimize_conversion_ed
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

prng = 3
random.seed(prng)
np.random.seed(prng)
    

# wb = load_workbook('Datafile for Simulation.xlsx') # Reading data file
# sheetname='Sobol_Data_real'
# ws1 = wb[sheetname]
# wb = load_workbook('experimental_data.xlsx') # Reading data file
# sheetname="U_P"
# wb = load_workbook('factrdata1.xlsx') # Reading data file
# sheetname="DoE_D_B_Eds"
# ws1 = wb[sheetname]
# wb = load_workbook('datasheet.xlsx') # Reading data file
# sheetname="MBDoE_BO_edited"
wb = load_workbook('datasheet.xlsx') # Reading data file
sheetname="Sheet50"
ws1 = wb[sheetname]
row_count = 0
for row in ws1:
    if not any([cell.value == None for cell in row]):
        row_count += 1

print(row_count)
n_prelim = row_count - 1
print("nprelim", n_prelim)
n_phi = 4 # number of design variables, excluding the outlet pressure
n_u = 5 # number of controls
n_y = 3 # number of measured responses
u_p = np.zeros([n_prelim,n_u])
y_meas = np.zeros([n_prelim,n_y])
Pin_meas = np.zeros(n_prelim)
    
for i in range(n_prelim):
    u_p[i,0] = ws1['B'+str(i + 2)].value
    u_p[i,1] = ws1['C'+str(i + 2)].value
    u_p[i,2] = ws1['E'+str(i + 2)].value
    u_p[i,3] = ws1['D'+str(i + 2)].value
    u_p[i,4] = ws1['F'+str(i + 2)].value
    y_meas[i,0] = ws1['I'+str(i + 2)].value
    y_meas[i,1] = ws1['H'+str(i + 2)].value
    y_meas[i,2] = ws1['G'+str(i + 2)].value
    Pin_meas[i] = ws1['J'+str(i + 2)].value
        
st = []
st += np.shape(u_p)[0] * [np.array([0, 0.01])]
    
ym = list(range(np.shape(u_p)[0]))
n_y = np.shape(y_meas)[1]
for i in range(np.shape(u_p)[0]):
    ym[i] = np.zeros([np.shape(st[i])[0],n_y])
ym_array = np.asarray(ym)
for i in range(np.shape(u_p)[0]):
    ym_array[i][0] = np.array([u_p[i][2],u_p[i][2]*u_p[i][3],0.0])
    ym_array[i][1] = y_meas[i]

# print('ymarray', ym_array)
print('ymeasure', y_meas)

lb_m1 = [0,0]
ub_m1 = [2e2,2e2]
ig_m1 = [6.9,7.3]

lb_m2 = [0,0,0,0,0,0]
ub_m2 = [2e2,2e2,2e2,2e2,2e2,2e2]
ig_m2 = [8.9,5.4,3.7,1.4,4.3,1.1]

lb_m3 = [0,0,0,0,0,0]
ub_m3 = [2e2,2e2,2e2,2e2,2e2,2e2]
ig_m3 = [2.0,9.2,5.6,3.5,10.6,9.0]
    
# NOTE: Measurement error covariance matrix
sigma_y = np.array([[0.00043**2, 0, 0], [0, 0.00202**2, 0], [0, 0, 0.00051**2]])
    
sigma_P = 0.005
    
pdtheta_hat = []
pdtheta_hat += [minimize_scalar(MLE_pd, bounds = (1e-6,1e6), method = 'bounded', 
                              args = (u_p,Pin_meas,sigma_P)).x]
    
alpha = 0.05
conf_level = 1.0 - alpha
y_cov = np.array([[0.00043**2, 0, 0], [0, 0.00202**2, 0], [0, 0, 0.00051**2]])
mexp =16
    

# NOTE: Parameter estimation of pressure drop model
Pi = []
Pi += [minimize_scalar(MLE_pd, bounds = (1e-6,1e6), method = 'bounded', 
                              args = (u_p[0:mexp],Pin_meas[0:mexp],sigma_P)).x]

print('u_p')
print (u_p)
print('u_p (1)(2)')
print(u_p[1][4])

# truetheta3 = [5.77, 6.72, 5.87, 9.51,10.17,7.98]
# truetheta3 = [6.1598, 8.0198, 3.97705,9.135,10.3558,6.3155]
# truetheta3 = [6.136949, 7.9217, 3.9411, 8.823,10.382,6.4295]
# truetheta3=[6.125, 7.7419, 4.212,10.5813,10.363,6.313]
truetheta3=[6.12403, 7.8283, 4.1813,10.3933,10.3662,6.339]
# truetheta3 = [6.08, 7.4486, 3.857,8.0864,10.4499,7.034]
outputs = []
for i, value in enumerate(u_p):
    result = insilico_exp(value, truetheta3, km3, pdtheta_hat[0], y_cov)
    outputs.append(result)


print('output results', outputs)
observedfish = obs_FIM(u_p,y_meas,km3,truetheta3, pdtheta_hat[0])
print("Overall Fisher", np.linalg.det(observedfish))



# Compute the Absolute Fisher Information for each experiment
absolute_fishers = [
    FIM_fun(u_p[i], y_meas, km3, truetheta3, pdtheta_hat[0]) for i in range(mexp)
]

# Compute the Absolute Fisher Information for each experiment
absolute_fishers = [
    FIM_fun(u_p[i], y_meas, km3, truetheta3, pdtheta_hat[0]) for i in range(mexp)
]

# Create a list to store the results
results = []

for idx, fisher_matrix in enumerate(absolute_fishers):
    # Calculate the determinant of the Fisher information matrix
    # Calculate the determinant, trace, and Frobenius norm
    determinant = np.linalg.det(fisher_matrix)
    determinant_trace = np.trace(fisher_matrix)
    determinant_frobenius_fim = np.linalg.norm(fisher_matrix)
    results.append({
        'Experiment No.': idx + 1,
        'Determinant': determinant,
        'Trace': determinant_trace,
        'Frobenius Norm': determinant_frobenius_fim
    })

# Convert the list of results into a DataFrame
df = pd.DataFrame(results)
# Save the DataFrame to an Excel file
output_file = "fisher_information_results.xlsx"
sheet_name = "FIM Results" + str(mexp)  # Specify the sheet name

# Use the ExcelWriter to specify the sheet name
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Results saved to {output_file} in the sheet '{sheet_name}'")    
# print(df)
    # print("The absolute fisher information for Experiment No.", idx+1)
    # determinant = np.linalg.det(fisher_matrix)
    # determinant_trace = np.trace(fisher_matrix)
    # determinant_frobenius_fim = np.linalg.norm(fisher_matrix)

    # print("original abs fisher", determinant)
    # print("original abs fisher_trace", determinant_trace)
    # print("original abs fisher_frobenius", determinant_frobenius_fim)

    # Ensure mexp does not exceed the length of the list

# Now use the adjusted determinants to calculate the observed Fisher
# Assuming obs_FIM is a function that computes observed Fisher based on the adjusted determinants
# determinants
u_p = u_p[:mexp]
y_meas = y_meas[:mexp, :]
observedfish = obs_FIM_ed(u_p, y_meas, km3, truetheta3, pdtheta_hat[0], adjusted_determinants=None)

# Print the new overall Fisher value
print(f"New Overall Fisher: {observedfish}")

observedfish = obs_FIM_ed(u_p,y_meas,km3,truetheta3, pdtheta_hat[0],adjusted_determinants=None)
print("New Overall Det Fisher", np.linalg.det(observedfish))



# NOTE: Calibration and posterior analysis of kinetic models
FIM1_obs, COV1_obs, COR1_obs, dof1, CI1, tval1 = [], [], [], [], [], []
FIM2_obs, COV2_obs, COR2_obs, dof2, CI2, tval2 = [], [], [], [], [], []
FIM3_obs, COV3_obs, COR3_obs, dof3, CI3, UCI, tval3 = [], [], [], [], [], [], []

# # # !!! One needs to put ipopt executable file in the path to use the solver
# # # NOTE: Parameter estimation using Pyomo model for power law kinetics
# current_directory = os.getcwd()
# # #solver_m1 = py.SolverFactory('ipopt', executable ='C:/Users/Mike/AnacondaProgram/Library/Ipopt-3.14.16-win64-msvs2019-md/bin/ipopt.exe')
# solver_m1 = py.SolverFactory('ipopt', executable =current_directory+ '/'+ 'Ipopt-3.14.16-win64-msvs2019-md/bin/ipopt.exe') #defining the solver

# # #solver_m1 = cyipopt.
# optmodel_m1 = optikm1(u_p[0:mexp], ym_array[0:mexp], st[0:mexp], Pi[-1], ig_m1, lb_m1, ub_m1) #define the optimisation model
# soln_m1 = solver_m1.solve(optmodel_m1) #defining the solver to solve the optimisation model
# est_m1 = np.array([py.value(optmodel_m1.theta[0]), py.value(optmodel_m1.theta[1])]) #obtaining the result...parameter estimated
# print('estimated parameters for Power Law',est_m1)
# objf_m1 = py.value(optmodel_m1.objfun) #get the chi-square which is the objective funtionc
# FIM1_obs += [obs_FIM(u_p[0:mexp],y_meas,km1,est_m1,Pi[-1])]
# COV1_obs += [np.linalg.inv(FIM1_obs[-1])]
# dof1 += [mexp * np.shape(y_meas)[1] - np.shape(est_m1)[0]]
# CI1 += [tvalue_fun(est_m1,COV1_obs[-1],dof1[-1])[0]]
# tval1 += [tvalue_fun(est_m1,COV1_obs[-1],dof1[-1])[1]] #returns a vector value.
# tref1 = stats.t.ppf((1-alpha),dof1[-1]) #reference t-value for model 1
# COR1_obs += [obs_COR(COV1_obs[-1])]
# sim1 = residual(u_p[0:mexp],est_m1,km1,y_meas[0:mexp],Pi[-1])[2]
# resid1 = residual(u_p[0:mexp],est_m1,km1,y_meas[0:mexp],Pi[-1])[0]
# nresid1 = residual(u_p[0:mexp],est_m1,km1,y_meas[0:mexp],Pi[-1])[1]
# pred_cov1 = np.zeros([mexp,np.shape(y_meas)[1]])
# for i in range(mexp):
#     pred_cov1[i] = pred_error(u_p[i],y_meas,km1,est_m1,COV1_obs[-1],Pi[-1],dof1[-1])[1]


# NOTE: Parameter estimation using Pyomo model for LHHW model
solver_m2 = py.SolverFactory('ipopt', executable ='C:/Users/Mike/AnacondaProgram/Library/Ipopt-3.14.16-win64-msvs2019-md/bin/ipopt.exe')
optmodel_m2 = optikm2(u_p[0:mexp], ym_array[0:mexp], st[0:mexp], Pi[-1], ig_m2, lb_m2, ub_m2)
soln_m2 = solver_m2.solve(optmodel_m2)
est_m2 = np.array([py.value(optmodel_m2.theta[0]), py.value(optmodel_m2.theta[1]), py.value(optmodel_m2.theta[2]), py.value(optmodel_m2.theta[3]), py.value(optmodel_m2.theta[4]), py.value(optmodel_m2.theta[5])])
objf_m2 = py.value(optmodel_m2.objfun)
FIM2_obs += [obs_FIM(u_p[0:mexp],y_meas,km2,est_m2,Pi[-1])]
COV2_obs += [np.linalg.inv(FIM2_obs[-1])]
dof2 += [mexp * np.shape(y_meas)[1] - np.shape(est_m2)[0]]
CI2 += [tvalue_fun(est_m2,COV2_obs[-1],dof2[-1])[0]]
tval2 += [tvalue_fun(est_m2,COV2_obs[-1],dof2[-1])[1]]
tref2 = stats.t.ppf((1-alpha),dof2[-1])
COR2_obs += [obs_COR(COV2_obs[-1])]
sim2 = residual(u_p[0:mexp],est_m2,km2,y_meas[0:mexp],Pi[-1])[2]
resid2 = residual(u_p[0:mexp],est_m2,km2,y_meas[0:mexp],Pi[-1])[0]
nresid2 = residual(u_p[0:mexp],est_m2,km2,y_meas[0:mexp],Pi[-1])[1]
pred_cov2 = np.zeros([mexp,np.shape(y_meas)[1]])
for i in range(mexp):
    pred_cov2[i] = pred_error(u_p[i],y_meas,km2,est_m2,COV2_obs[-1],Pi[-1],dof2[-1])[1]
    
    
# NOTE: Parameter estimation using Pyomo model for MVK model
solver_m3 = py.SolverFactory('ipopt', executable ='C:/Users/Mike/AnacondaProgram/Library/Ipopt-3.14.16-win64-msvs2019-md/bin/ipopt.exe')
optmodel_m3 = optikm3(u_p[0:mexp], ym_array[0:mexp], st[0:mexp], Pi[-1], ig_m3, lb_m3, ub_m3) 
# I had to format this part
soln_m3 = solver_m3.solve(optmodel_m3)
est_m3 = np.array([py.value(optmodel_m3.theta[0]), py.value(optmodel_m3.theta[1]), py.value(optmodel_m3.theta[2]), py.value(optmodel_m3.theta[3]), py.value(optmodel_m3.theta[4]), py.value(optmodel_m3.theta[5])])
objf_m3 = py.value(optmodel_m3.objfun) #objf_m3  this is the chi-squre value for the model...
print('Print Chi-Square value', objf_m3)
FIM3_obs += [obs_FIM(u_p[0:mexp],y_meas,km3,est_m3,Pi[-1])]
print('Fisher Total', np.linalg.det(FIM3_obs))
COV3_obs += [np.linalg.inv(FIM3_obs[-1])]
# print('covariance', COV3_obs)
print('Fisher Information Matrix', FIM3_obs)
print('Fisher Total', np.linalg.det(FIM3_obs))
print('diagonal of the matrix', np.diagonal(COV3_obs))
print('parameter_estimates', est_m3)
dof3 += [mexp * np.shape(y_meas)[1] - np.shape(est_m3)[0]]
CI3 += [tvalue_fun(est_m3,COV3_obs[-1],dof3[-1])[0]]
print("confidence interval", CI3)
# Calculate uncertainty (half-width of each confidence interval)
UCI += [ci / 2 for ci in CI3[-1]]
print("uncertainty", UCI)
tval3 += [tvalue_fun(est_m3,COV3_obs[-1],dof3[-1])[1]]  # t-test for the model 3.
print("t-test result", tval3)
COR3_obs += [obs_COR(COV3_obs[-1])]
sim3 = residual(u_p[0:mexp],est_m3,km3,y_meas[0:mexp],Pi[-1])[2]
sim31 = np.zeros([mexp,np.shape(y_meas)[1]])
for i in range(mexp):
    sim31[i] = pred_error(u_p[i],y_meas,km3,est_m3,COV3_obs[-1],Pi[-1],dof3[-1])[0]
resid3 = residual(u_p[0:mexp],est_m3,km3,y_meas[0:mexp],Pi[-1])[0]
nresid3 = residual(u_p[0:mexp],est_m3,km3,y_meas[0:mexp],Pi[-1])[1]
tref3 = stats.t.ppf((1-alpha),dof3[-1])  #reference t - value
pred_cov3 = np.zeros([mexp,np.shape(y_meas)[1]])
for i in range(mexp):
    pred_cov3[i] = pred_error(u_p[i],y_meas,km3,est_m3,COV3_obs[-1],Pi[-1],dof3[-1])[1]
    print(pred_cov3[i])

# # NOTE: Chi-square goodness-of-fit test
# refchisq1 = stats.chi2.ppf((conf_level),dof1[-1]) #reference chisquare value for model 1
# refchisq2 = stats.chi2.ppf((conf_level),dof2[-1])#reference chisquare value for model 2
# refchisq3 = stats.chi2.ppf((conf_level),dof3[-1]) #reference chisquare value for model 3

# Step 1: Extract parameter estimates and t-values
parameter_estimates = est_m3  # Parameter estimates
t_values = tval3[-1]  # Corresponding t-values (last one in tval3 list)
reference_t_value = tref3  # The reference t-value for comparison

# Step 2: Create a list of 'pass/fail' for each t-test
t_test_results = ['Pass' if abs(t) > reference_t_value else 'Fail' for t in t_values]

# Step 3: Create a DataFrame to show the table
df = pd.DataFrame({
    'Parameter Estimate': parameter_estimates,
    't-value': t_values,
    'Reference t-value': [reference_t_value] * len(parameter_estimates),  # Same ref value for all
    'Pass/Fail': t_test_results
})

# Step 4: Display the table
print(df.to_string(index=False))


# # NOTE: Computation of probability of model correctness
# obspr1 = []
# obspr1 += [mprob1([objf_m1,objf_m2,objf_m3],[dof1[-1],dof2[-1],dof3[-1]])[1]]
    
# obspr2 = []
# obspr2 += [mprob2([objf_m1,objf_m2,objf_m3],[dof1[-1],dof2[-1],dof3[-1]])]

n_dexp = 1 # Number of experiments to be designed using MBDoE method

# # NOTE: MBDoE for model discrimination (between model 2 and 3)
# ig_md1, b_md1 = initialisation1_mbdoemd_BF(y_meas,km2,km3,est_m2,est_m3,FIM2_obs[-1],FIM3_obs[-1],Pi[-1],n_dexp,n_u)#,prng)
# ig_md0, b_md0 = initialisation0_mbdoemd_BF(y_meas,km2,km3,est_m2,est_m3,FIM2_obs[-1],FIM3_obs[-1],Pi[-1],n_dexp,n_u,prng)
# sol_md1 = minimize(mbdoemd_BF,ig_md1,method = 'SLSQP', bounds = (b_md1), args = (y_meas,km2,km3,est_m2,est_m3,FIM2_obs[-1],FIM3_obs[-1],Pi[-1],n_dexp,n_phi))
# sol_md0 = minimize(mbdoemd_BF,ig_md0,method = 'SLSQP', bounds = (b_md0), args = (y_meas,km2,km3,est_m2,est_m3,FIM2_obs[-1],FIM3_obs[-1],Pi[-1],n_dexp,n_phi))
# print("Inital gueses", ig_md0)
# print("b_mdo", b_md0)

ig_pp1, b_pp1 = initialisation1_mbdoepp(y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_u)#,prng)
# ig_pp0, b_pp0 = initialisation0_mbdoepp(y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_u,prng)
sol_pp1 = minimize(mbdoepp, ig_pp1, method = 'SLSQP', bounds = (b_pp1), args = (y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_phi))
# sol_pp0 = minimize(mbdoepp, ig_pp0, method = 'SLSQP', bounds = (b_pp0), args = (y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_phi))


# sol_conversion = minimize(optimize_conversion, ig_md0, method = 'SLSQP', bounds = b_md0, args = (y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_phi))

print("conditions", u_p[10])
initial_conditions = np.array(u_p[10]) 

# import numpy as np
# from scipy.optimize import minimize

# # Define objective function to maximize FIM's sensitivity for the fifth parameter
# def fim_objective(conditions, est_params, y_meas, Pi):
#     # Adjust experimental conditions based on input `conditions`
#     temp, massflow, moleratio, inletconc, pinlet = conditions
    
#     # Wrap conditions in a list to make it compatible with `obs_FIM` expected input format
#     FIM_current = obs_FIM([[temp, massflow, moleratio, inletconc, pinlet]], y_meas, km3, est_params, Pi)
    
#     # Check determinant to see if the matrix is nearly singular
#     det_FIM = np.linalg.det(FIM_current)
#     if np.abs(det_FIM) < 1e-10:
#         print("Warning: FIM is nearly singular, adding regularization.")
#         # Apply regularization
#         FIM_current += np.eye(FIM_current.shape[0]) * 1e-6

#     # Now invert the matrix (regularized version)
#     COV_current = np.linalg.inv(FIM_current)
#     variance_fifth_param = COV_current[4, 4]  # fifth parameter variance
    
#     # We want to minimize variance, so return it directly as the objective
#     return variance_fifth_param

# # Set bounds for experimental conditions to ensure practical values
# bounds = [
#     (250, 355),  # Temperature bounds
#     (20, 30),    # Massflow bounds
#     (2, 4),      # Moleratio bounds
#     (0.005, 0.025),  # Inlet concentration bounds
#     (1.5, 2.0)   # Pinlet bounds
# ]

# # Initial guess for experimental conditions

# initial_conditions = np.array(u_p[mexp-1])  # Ensure this is a single 1D array

# # Run optimization to find conditions that minimize the uncertainty in the fifth parameter
# result = minimize(fim_objective, initial_conditions, args=(est_m3, y_meas, Pi[-1]), bounds=bounds)

# # Extract optimized conditions
# optimized_conditions = result.x
# print("Optimized Conditions:", optimized_conditions)



# sol_conversion = minimize(optimize_conversion, ig_md0, method = 'SLSQP', bounds = b_md0, args = (y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_phi))
# # NOTE: MBDoE for parameter precision of model 3
# print('solution_conversion', sol_conversion.x)
# op_cond=sol_conversion.x
# op_cond=sol_conversion.x

print('solution_conversion', sol_pp1.x)
op_cond=sol_pp1.x
op_cond=sol_pp1.x

# conversion, uncertainty = optimize_conversion_ed(op_cond, y_meas, km3, est_m3, FIM3_obs[-1], Pi[-1], n_dexp, n_phi)

# Now you can print both values
# print(f"Optimal Conversion: {conversion:.4f}%, Uncertainty: ±{uncertainty:.4f}%")

print('operating conditions', sol_pp1.x)
op_cond=sol_pp1.x
op_cond = np.array(op_cond)
# op_cond=np.array([303.0,	30,	4,	0.0017])

# Additional value to add
Pinmodel = 1.2973272


# Step 1: Add Pinmodel to the operating_conditions array
op_cond = np.append(op_cond, Pinmodel)
print(op_cond)
# Step 2: Reshape to a column vector (5 rows, 1 column)
op_cond = op_cond.reshape(1, -1)
outputs = []
for i, value in enumerate(op_cond):
    result = insilico_exp(value, truetheta3, km3, pdtheta_hat[0], y_cov)
    outputs.append(result)
# ig_pp1, b_pp1 = initialisation1_mbdoepp(y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_u)#,prng)
# ig_pp0, b_pp0 = initialisation0_mbdoepp(y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_u,prng)
# sol_pp1 = minimize(mbdoepp, ig_pp1, method = 'SLSQP', bounds = (b_pp1), args = (y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_phi))
# sol_pp0 = minimize(mbdoepp, ig_pp0, method = 'SLSQP', bounds = (b_pp0), args = (y_meas,km3,est_m3,FIM3_obs[-1],Pi[-1],n_dexp,n_phi))

print("Solution_guess", sol_pp1.x)
print("My Results", sol_pp1)
print("Simulated Results for Arun", outputs)

initial_data = np.array(u_p)
response_data = np.array(y_meas)
print("Length of U_P",len(u_p))
print("Length of YMeas",len(y_meas))
# Ensure that both arrays have the same number of rows
if u_p.shape[0] != y_meas.shape[0]:
    raise ValueError("u_p and ymeas must have the same number of rows.")

# Concatenate u_p and ymeas horizontally (axis=1)
# combined_data = np.hstack((initial_data, response_data))
# # Operating conditions and simulated results
# operating_conditions = np.array([3.27291300e+02, 2.24409757e+01, 5.00000000e-03, 3.84285092e+00])
# simulated_results = np.array([-0.00022083,  0.01142536,  0.00459282])  # CH4, O2, CO2

operating_conditions=np.array(op_cond)
simulated_results=np.array(outputs)

print("This is the initial data", initial_data)

print("---------------------------------------")

print("This is the response", y_meas)


# print("This below is the merged inital data",combined_data)
# Define POutlet and PInlet for the first 12 experiments
POutlet = np.array([1.3062266, 1.3017769, 1.3062266, 1.3062266, 1.2884278, 1.3062266, 
                    1.3062266, 1.3062266, 1.3017769, 1.3062266, 1.3017769, 1.3017769])

PInlet = np.array([1.6619685, 1.7553747, 1.6530726, 1.7598226, 1.6619685, 1.7608226, 
                   1.8087497, 1.951083, 1.8265414, 1.9599788, 1.8131976, 1.9777705])

# Experimental data for the first 12 experiments (Replace this with actual values)
U_P_data = {
    'Experiment': [1,2,3,4,5,6,7, 8, 9, 10, 11, 12],
    'Temperature': [320, 310, 315, 325, 330, 335, 340, 345, 350, 355, 360, 365],  # Replace with actual values
    'Mass Flow': [20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31],                # Replace with actual values
    'Mole Ratio': [0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.11, 0.12],  # Replace
    'Inlet Concentration': [4, 4.1, 4.2, 4.3, 4.4, 4.5, 4.6, 4.7, 4.8, 4.9, 5.0, 5.1],      # Replace
    'Poutlet': POutlet,
    'yCO2': [0.1, 0.12, 0.13, 0.11, 0.15, 0.14, 0.13, 0.16, 0.17, 0.18, 0.19, 0.2],         # Replace
    'yO2': [0.21, 0.22, 0.23, 0.24, 0.25, 0.26, 0.27, 0.28, 0.29, 0.3, 0.31, 0.32],         # Replace
    'yCH4': [0.02, 0.018, 0.019, 0.02, 0.021, 0.022, 0.023, 0.024, 0.025, 0.026, 0.027, 0.028], # Replace
    'Pinlet': PInlet
            }

# Define the headers (they stay constant)
headers = ['Experiment', 'Temperature', 'Mass Flow', 'Mole Ratio', 'Inlet Concentration', 'Poutlet', 'yCO2', 'yO2', 'yCH4', 'Pinlet']
# Create the U_P DataFrame with the actual experimental data
# Function to update u_p dynamically
# Merging headers with values using zip()
# merged_data = dict(zip(headers, combined_data))


# U_P = pd.DataFrame(merged_data)

print("operating_conditions", operating_conditions)
print("simulated_Results",simulated_results)
# Concatenate the operating conditions and simulated results to form SimU
# Flatten both arrays into 1D arrays
operating_conditions_flat = operating_conditions.flatten()
simulated_results_flat = simulated_results.flatten()

# Concatenate the flattened arrays along axis 0
SimU = np.concatenate((operating_conditions_flat, simulated_results_flat))

print("simulated results_real", SimU)

# # Excel file path
# excel_file = 'experiment_real.xlsx'

# # Function to get the current row count from Excel
# def get_row_count(file_name, sheet_name='U_P'):
#     try:
#         # Load the workbook and get the active sheet
#         wb = load_workbook(file_name)
#         sheet = wb[sheet_name]
#         return sheet.max_row  # Returns the last row with data
#     except FileNotFoundError:
#         return 1  # If file doesn't exist, start at experiment 1

# # Get the next experiment number based on current row count
# current_row_count = get_row_count(excel_file)
# experiment_number = current_row_count - 1  # Start from the next available row

# # Create a SimU dataframe (1 row, matching U_P columns)
# SimU_df = pd.DataFrame([{
#     'Experiment': experiment_number,
#     'Temperature': SimU[0],
#     'Mass Flow': SimU[1],
#     'Mole Ratio': SimU[2],
#     'Inlet Concentration': SimU[3],
#     'POutlet': np.random.choice(POutlet),  # Example value (adjust with real value)
#     'yCO2': SimU[7],
#     'yO2': SimU[6],
#     'yCH4': SimU[5],
#     'PInlet': np.random.choice(PInlet)  # Example value (adjust with real value)
# }])

# # Define the Excel file path
# excel_file = 'experiment_real.xlsx'

# # Write U_P to Excel sheet with headers and index
# with pd.ExcelWriter(excel_file, engine='openpyxl', mode='w') as writer:
#     U_P_data.to_excel(writer, sheet_name='U_P', index=False)

# # Function to append new SimU row to Excel sheet
# def append_to_excel(new_row_df, file_name, sheet_name='U_P'):
#     # Append SimU row to the existing Excel sheet
#     with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#         new_row_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row)

# # Every time SimU is updated, append it to the Excel sheet
# append_to_excel(SimU_df, excel_file)

print('simulated exp results', outputs)


end_time = time.time()
print(f"Total execution time: {end_time - start_time:.2f} seconds")